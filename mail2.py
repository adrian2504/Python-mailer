
import smtplib
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders


email_address="mail id" #sender email
password="password" #sender password

def send_email(i,subject,msg,User_Email):
    email = User_Email
    print(email)
    print(email_address)
    try:
        server = smtplib.SMTP('smtp.gmail.com',587)
        server.ehlo()
        server.starttls()
        server.login(email_address,password)
        message = 'Subject: {}\n\n{}'.format(subject,msg)
        server.sendmail(email_address,email, message)

        print(i+" success "+User_Email)
        server.close()
    except :
        print(" Failed to send email to ")



data = openpyxl.load_workbook('excel sheet')
sheet = data.active

i = 2
cell = 1
while cell != 0:

    name = sheet.cell(row=i, column=1)
    email = sheet.cell(row=i, column=2)
    name = name.value
    email = email.value
    print(name + email)
    if name == None and email == None:
        cell = 0
        break
    subject = "subject message here"  # mail subject
    msg = "body message here "+ name #mail msg
    send_email(i,subject, msg, email)  #sending mail
    i=i+1


