import smtplib
import imghdr
import openpyxl
from email.message import EmailMessage





Sender_Email="mention senders mail" #sender email
Password="senders password here" #sender password

def send_email(i,subject,msg,User_Email, name):
    Reciever_Email = User_Email
    print(Reciever_Email)

    newMessage = EmailMessage()
    newMessage['Subject'] = subject
    newMessage['From'] = Sender_Email
    newMessage['To'] = Reciever_Email
    newMessage.set_content(msg + ' Please check below for participation certificate attachment')
    with open("../result-image/certificate-"+name+'.png', 'rb') as f:
        image_data = f.read()
        image_type = imghdr.what(f.name)
        image_name = f.name
    newMessage.add_attachment(image_data, maintype='image', subtype=image_type, filename=image_name)
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(Sender_Email, Password)
        smtp.send_message(newMessage)





data = openpyxl.load_workbook('mention the excel file here')
sheet = data.active

i = 2
cell = 1
while cell != 0:

    name = sheet.cell(row=i, column=1)
    email = sheet.cell(row=i, column=2)
    name = name.value
    email = email.value
    if name == None and email == None:
        cell = 0
        break
    print(name + email)
    subject = "Write subject here"  # mail subject
    msg = "We hope you "+ name + " mention message of body here" #mail msg
    send_email(i,subject, msg, email, name)  #sending mail
    i=i+1